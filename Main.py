import cv2
import os
import numpy as np
import csv
import tkinter as tk
from tkinter import Tk, Entry, Label, Button
from tkinter.filedialog import askopenfilename
import pandas as pd
import pyttsx3
from datetime import datetime
import openpyxl
from PIL import Image, ImageTk
from tkinter import simpledialog
import shutil

engine = pyttsx3.init()


def display_image(image):
    cv2.imshow('Training Image', image)
    cv2.waitKey(100)
    cv2.destroyAllWindows()

def create_dataset_with_webcam():
    def get_input():
        name = name_entry.get()
        roll_number = roll_number_entry.get()
        root.destroy()
        return name, roll_number

    root = Tk()
    root.title("Enter Details")

    name_label = Label(root, text="Name:")
    name_label.grid(row=0, column=0, padx=10, pady=5, sticky="e")
    name_entry = Entry(root)
    name_entry.grid(row=0, column=1, padx=10, pady=5)

    roll_number_label = Label(root, text="Roll Number:")
    roll_number_label.grid(row=1, column=0, padx=10, pady=5, sticky="e")
    roll_number_entry = Entry(root)
    roll_number_entry.grid(row=1, column=1, padx=10, pady=5)

    confirm_button = Button(root, text="Confirm", command=get_input)
    confirm_button.grid(row=2, columnspan=2, pady=10)

    root.mainloop()

    name, roll_number = get_input()

    directory = f"dataset/{name}_{roll_number}"
    if not os.path.exists(directory):
        os.makedirs(directory)

    cap = cv2.VideoCapture(0)
    face_cascade = cv2.CascadeClassifier(
        cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')  

    count = 0
    speak("Please look at the webcam. The system is capturing your images.")
    while True:
        ret, frame = cap.read()  
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)  
        faces = face_cascade.detectMultiScale(gray, 1.3, 5) 

        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
            roi_gray = gray[y:y + h, x:x + w]
            roi_color = frame[y:y + h, x:x + w]
            file_path = f"{directory}/img_{count}.jpg"
            cv2.imwrite(file_path, roi_color) 
            display_image(roi_color)  
            count += 1

        cv2.imshow('Create Dataset', frame)  
        if cv2.waitKey(1) & 0xFF == ord('q') or count >= 100:  
            break

    cap.release() 
    cv2.destroyAllWindows()

    speak("Dataset creation completed successfully.")  

def train_existing_dataset():
    if os.path.exists("trained_model.yml"):  #
        recognizer = cv2.face.LBPHFaceRecognizer_create()  
        recognizer.read("trained_model.yml") 
        speak("Existing model loaded.")  
    else:
        recognizer = cv2.face.LBPHFaceRecognizer_create()  
        speak("New model created.")  

    detector = cv2.CascadeClassifier(
        cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')  

    faces, ids = get_images_and_labels("dataset")  
    for face in faces:
        display_image(face)  
    recognizer.update(faces, np.array(ids))  
    recognizer.save("trained_model.yml")  
    speak("Model updated successfully")  


def recognize_from_webcam():
    recognizer = cv2.face.LBPHFaceRecognizer_create()  
    recognizer.read("trained_model.yml")  
    face_cascade = cv2.CascadeClassifier(
        cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')  
    font = cv2.FONT_HERSHEY_SIMPLEX  

    cap = cv2.VideoCapture(0)  

    # Get today's date
    today_date = datetime.now().strftime("%Y-%m-%d")
    wb_name = f"attendance_{today_date}.xlsx"  

    try:
        if not os.path.exists(wb_name):
            wb = openpyxl.Workbook()  
            ws = wb.active
            ws.title = today_date
            ws.append(["Roll Number", "Name", "Date", "Time", "Duration"])  # Append headers to worksheet
        else:
            wb = openpyxl.load_workbook(wb_name)  # Load existing workbook
            if today_date not in wb.sheetnames:
                ws = wb.create_sheet(title=today_date)  # Create new worksheet for today's date
                ws.append(["Roll Number", "Name", "Date", "Time", "Duration"])  # Append headers to worksheet
            else:
                ws = wb[today_date]  # Use existing worksheet for today's date

        start_time = datetime.now()  # Record start time of recognition process outside the loop

        while True:
            ret, frame = cap.read()  # Read frame from webcam
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)  # Convert frame to grayscale
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)  # Detect faces in the grayscale frame

            for (x, y, w, h) in faces:
                roi_gray = gray[y:y + h, x:x + w]  # Extract region of interest (ROI) from grayscale image
                id_, confidence = recognizer.predict(roi_gray)  # Recognize face and get ID and confidence level
                if confidence < 70:
                    # Fetch name and roll number from dataset based on ID
                    label = get_label(id_)
                    end_time = datetime.now()  # Record end time of recognition process
                    duration = end_time - start_time  # Calculate duration of recognition process

                    # Format duration in hours:minutes:seconds format
                    duration_str = str(duration).split(".")[0]  # Remove milliseconds
                    cv2.putText(frame, duration_str, (x, y - 20), font, 0.8, (0, 255, 0), 2, cv2.LINE_AA)  # Overlay duration as label
                    mark_attendance(ws, label, id_, duration)  # Mark attendance for recognized person with duration
                else:
                    label = "Unknown"

                cv2.putText(frame, label, (x, y), font, 1, (0, 255, 0), 2, cv2.LINE_AA)  # Overlay label on frame
                cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)  # Draw rectangle around detected face

            cv2.imshow('Recognize from Webcam', frame)  # Display frame with overlaid text and rectangles
            if cv2.waitKey(1) & 0xFF == ord('q'):  # Exit loop if 'q' key is pressed
                break

        cap.release()  # Release the webcam
        cv2.destroyAllWindows()  # Close all OpenCV windows
        wb.save(wb_name)  # Save workbook to file
    except PermissionError:
        print(f"Permission error: Unable to save the file '{wb_name}'. Make sure you have permission to write to the directory.")

# Function to get images and labels from dataset directory
def get_images_and_labels(path):
    image_paths = [os.path.join(path, person) for person in os.listdir(path)]  # Get list of image paths
    faces = []
    ids = []

    for image_path in image_paths:
        person_id = os.path.basename(image_path).split('_')[1]  # Extract person's ID from image path
        for image_file in os.listdir(image_path):
            image = cv2.imread(os.path.join(image_path, image_file), cv2.IMREAD_GRAYSCALE)  # Read image as grayscale
            faces.append(image)  # Append image to faces list
            ids.append(int(person_id))  # Append ID to ids list

    return faces, ids


# Function to get label (name and roll number) based on ID
def get_label(id_):
    with open('students.csv', mode='r') as file:
        reader = csv.DictReader(file)  # Create CSV reader object
        for row in reader:
            if int(row['ID']) == id_:
                return f"{row['Name']} ({row['RollNumber']})"  # Return name and roll number based on ID
    return "Unknown"  # Return 'Unknown' if ID not found


# Function to extract faces from a photo
def extract_faces_from_photo():
    def confirm_accept_face(face):
        nonlocal accepted_faces
        choice = simpledialog.askstring("Accept Face", f"Accept this face? (y/n)\n\nPress 'y' to accept or 'n' to reject")
        if choice and choice.lower() == 'y':
            name = simpledialog.askstring("Enter Name", f"Enter name for the accepted face:")
            roll_number = simpledialog.askstring("Enter Roll Number", f"Enter roll number for the accepted face:")
            if name and roll_number:
                directory = f"dataset/{name}_{roll_number}"
                if not os.path.exists(directory):
                    os.makedirs(directory)
                for j in range(100):
                    cv2.imwrite(f"{directory}/img_{j}.jpg", face)
                accepted_faces += 1

    Tk().withdraw()  # Hide the small Tkinter window
    filename = askopenfilename()  # Open file dialog to select image file

    if filename:
        photo = cv2.imread(filename)  # Read photo from file
        face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')  # Load pre-trained face cascade classifier
        faces = face_cascade.detectMultiScale(cv2.cvtColor(photo, cv2.COLOR_BGR2GRAY), 1.3, 5)  # Detect faces in photo

        # Create the dataset directory if it doesn't exist
        if not os.path.exists("dataset"):
            os.makedirs("dataset")

        accepted_faces = 0
        speak("Faces are being extracted from the photo.")  # Speak message indicating face extraction process
        # Process each detected face
        for i, (x, y, w, h) in enumerate(faces):
            face = photo[y:y + h, x:x + w]  # Crop face from photo
            resized_face = cv2.resize(face, (400, 400))  # Resize face to standard size
            cv2.imshow(f"Face {i + 1}", resized_face)  # Display the cropped face
            cv2.waitKey(0)  # Wait for key press
            confirm_accept_face(face)
            # Close the face window
            cv2.destroyAllWindows()
        
        # Show a message box with the total number of accepted faces
        simpledialog.messagebox.showinfo("Accepted Faces", f"{accepted_faces} faces were accepted and added to the dataset.")

# Function to recognize faces from an uploaded image
def recognize_from_uploaded_image():
    Tk().withdraw()  # Hide the small Tkinter window
    filename = askopenfilename()  # Open file dialog to select image file

    if filename:
        photo = cv2.imread(filename)  # Read uploaded image
        if photo is not None:
            recognizer = cv2.face.LBPHFaceRecognizer_create()  # Create LBPH face recognizer object
            recognizer.read("trained_model.yml")  # Load trained model
            face_cascade = cv2.CascadeClassifier(
                cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')  # Load pre-trained face cascade classifier
            font = cv2.FONT_HERSHEY_SIMPLEX  # Set font for text overlay

            gray = cv2.cvtColor(photo, cv2.COLOR_BGR2GRAY)  # Convert image to grayscale
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)  # Detect faces in the grayscale image

            # Get today's date
            today_date = datetime.now().strftime("%Y-%m-%d")
            wb_name = f"attendance_{today_date}.xlsx"  # Define Excel file name based on date

            try:
                # Check if the workbook already exists, if not create a new one
                if not os.path.exists(wb_name):
                    wb = openpyxl.Workbook()  # Create new workbook object
                    ws = wb.active
                    ws.title = today_date
                    ws.append(["Roll Number", "Name", "Date", "Time"])  # Append headers to worksheet
                else:
                    wb = openpyxl.load_workbook(wb_name)  # Load existing workbook
                    if today_date not in wb.sheetnames:
                        ws = wb.create_sheet(title=today_date)  # Create new worksheet for today's date
                        ws.append(["Roll Number", "Name", "Date", "Time"])  # Append headers to worksheet
                    else:
                        ws = wb[today_date]  # Use existing worksheet for today's date

                for (x, y, w, h) in faces:
                    roi_gray = gray[y:y + h, x:x + w]  # Extract region of interest (ROI) from grayscale image
                    id_, confidence = recognizer.predict(roi_gray)  # Recognize face and get ID and confidence level
                    if confidence < 70:
                        # Fetch name and roll number from dataset based on ID
                        label = get_label(id_)
                        mark_attendance(ws, label, id_)  # Mark attendance for recognized person
                    else:
                        label = "Unknown"

                    cv2.putText(photo, label, (x, y), font, 1, (0, 255, 0), 2, cv2.LINE_AA)  # Overlay label on image
                    cv2.rectangle(photo, (x, y), (x + w, y + h), (255, 0, 0), 2)  # Draw rectangle around detected face

                wb.save(wb_name)  # Save workbook to file
                cv2.imshow('Recognize from Uploaded Image', photo)  # Display image with overlaid text and rectangles
                cv2.waitKey(0)  # Wait for key press
                cv2.destroyAllWindows()  # Close all OpenCV windows
            except PermissionError:
                print(f"Permission error: Unable to save the file '{wb_name}'. Make sure you have permission to write to the directory.")
        else:
            print("Error: Could not load image file.")  # Print error message if image file cannot be loaded
    else:
        print("No file selected.")  # Print message if no file is selected

def delete_face_from_dataset():
    name = simpledialog.askstring("Enter Name", "Enter the name of the person whose face you want to delete:")
    roll_number = simpledialog.askstring("Enter Roll Number", "Enter the roll number of the person whose face you want to delete:")
    if name and roll_number:
        # Construct the directory path based on name and roll number
        directory = f"dataset/{name}_{roll_number}"
        # Check if the directory exists
        if os.path.exists(directory):
            # Delete the directory and all its contents (faces)
            shutil.rmtree(directory)
            speak("Face deleted successfully.")
        else:
            speak("Face not found in the dataset.")
    else:
        speak("Invalid input. Please provide both name and roll number.")


# Function to mark attendance in Excel worksheet
def mark_attendance(ws, name, roll_number, duration=None):
    now = datetime.now()  # Get current date and time
    date_string = now.strftime("%Y-%m-%d")  # Format current date as string
    time_string = now.strftime("%H:%M:%S")  # Format current time as string

    # Check if the attendance for this person on this date has already been marked
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value == roll_number:
            existing_duration = row[4].value if len(row) > 4 else None  # Get existing duration from the worksheet if available
            if existing_duration is not None:
                existing_duration = datetime.strptime(existing_duration, "%H:%M:%S")  # Convert existing duration to datetime object
                if duration:
                    duration += existing_duration.time()  
            ws.delete_rows(row[0].row)  
            break

    if duration:
        ws.append([roll_number, name, date_string, time_string, str(duration)])  
    else:
        ws.append([roll_number, name, date_string, time_string])  
    print("Attendance marked successfully.")  

def speak(message):
    engine.say(message)  
    engine.runAndWait()  


# Main function to control program flow
def display_main_menu():
    root = tk.Tk()
    root.title("Smart Attendance System")

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    bg_image = Image.open("image.png")
    bg_image = bg_image.resize((screen_width, screen_height), Image.ANTIALIAS)
    bg_image = ImageTk.PhotoImage(bg_image)

    background_label = tk.Label(root, image=bg_image)
    background_label.place(x=0, y=0, relwidth=1, relheight=1)  # Fill the entire window
    background_label.image = bg_image  

    bg_color = "#3498db" 
    main_heading_bg_color = "#5DADE2"  
    btn_bg_color = "#1F618D"  
    hover_color = "#21618C"  

    canvas = tk.Canvas(root, width=800, height=600)
    canvas.pack(fill="both", expand=True)
    canvas.create_image(0, 0, image=bg_image, anchor="nw")

    label = tk.Label(canvas, text="Welcome to Smart Attendance Management System", font=("Helvetica", 24), bg=bg_color, fg="white")
    label.place(relx=0.5, rely=0.1, anchor="center")

    btn_create_dataset = tk.Button(canvas, text="Create Dataset with Webcam", command=create_dataset_with_webcam, bg=btn_bg_color, fg="white", font=("Helvetica", 16))
    btn_create_dataset.place(relx=0.25, rely=0.35, anchor="center")
    btn_create_dataset.bind("<Enter>", lambda event, btn=btn_create_dataset: btn.config(bg=hover_color))
    btn_create_dataset.bind("<Leave>", lambda event, btn=btn_create_dataset: btn.config(bg=btn_bg_color))

    btn_recognize_from_webcam = tk.Button(canvas, text="Recognize from Webcam", command=recognize_from_webcam, bg=btn_bg_color, fg="white", font=("Helvetica", 16))
    btn_recognize_from_webcam.place(relx=0.75, rely=0.35, anchor="center")
    btn_recognize_from_webcam.bind("<Enter>", lambda event, btn=btn_recognize_from_webcam: btn.config(bg=hover_color))
    btn_recognize_from_webcam.bind("<Leave>", lambda event, btn=btn_recognize_from_webcam: btn.config(bg=btn_bg_color))

    btn_extract_faces_from_photo = tk.Button(canvas, text="Extract Faces from Photo", command=extract_faces_from_photo, bg=btn_bg_color, fg="white", font=("Helvetica", 16))
    btn_extract_faces_from_photo.place(relx=0.25, rely=0.55, anchor="center")
    btn_extract_faces_from_photo.bind("<Enter>", lambda event, btn=btn_extract_faces_from_photo: btn.config(bg=hover_color))
    btn_extract_faces_from_photo.bind("<Leave>", lambda event, btn=btn_extract_faces_from_photo: btn.config(bg=btn_bg_color))

    btn_recognize_from_uploaded_image = tk.Button(canvas, text="Recognize from Uploaded Image", command=recognize_from_uploaded_image, bg=btn_bg_color, fg="white", font=("Helvetica", 16))
    btn_recognize_from_uploaded_image.place(relx=0.75, rely=0.55, anchor="center")
    btn_recognize_from_uploaded_image.bind("<Enter>", lambda event, btn=btn_recognize_from_uploaded_image: btn.config(bg=hover_color))
    btn_recognize_from_uploaded_image.bind("<Leave>", lambda event, btn=btn_recognize_from_uploaded_image: btn.config(bg=btn_bg_color))

    btn_train_existing_dataset = tk.Button(canvas, text="Train Existing Dataset", command=train_existing_dataset, bg=btn_bg_color, fg="white", font=("Helvetica", 16))
    btn_train_existing_dataset.place(relx=0.25, rely=0.75, anchor="center")
    btn_train_existing_dataset.bind("<Enter>", lambda event, btn=btn_train_existing_dataset: btn.config(bg=hover_color))
    btn_train_existing_dataset.bind("<Leave>", lambda event, btn=btn_train_existing_dataset: btn.config(bg=btn_bg_color))

    btn_delete_face_from_dataset = tk.Button(canvas, text="Delete Face from Dataset", command=delete_face_from_dataset, bg=btn_bg_color, fg="white", font=("Helvetica", 16))
    btn_delete_face_from_dataset.place(relx=0.75, rely=0.75, anchor="center")
    btn_delete_face_from_dataset.bind("<Enter>", lambda event, btn=btn_delete_face_from_dataset: btn.config(bg=hover_color))
    btn_delete_face_from_dataset.bind("<Leave>", lambda event, btn=btn_delete_face_from_dataset: btn.config(bg=btn_bg_color))

    # btn_mark_attendance = tk.Button(root, text="Mark Attendance", command=mark_attendance)
    # btn_mark_attendance.pack(pady=5)

    # Run the GUI
    root.mainloop()

# Call the function to display the main menu
display_main_menu()